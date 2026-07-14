from datetime import datetime
from pathlib import Path
from xml.etree.ElementTree import Element, SubElement, ElementTree, indent

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


BRAND_ORANGE = colors.HexColor("#F29129")
BRAND_DARK = colors.HexColor("#29292E")
BRAND_MUTED = colors.HexColor("#66666B")
BRAND_LINE = colors.HexColor("#D8D3CA")
BRAND_LIGHT = colors.HexColor("#FAF9F7")
LOGO_PATH = Path(__file__).resolve().parents[1] / "panel" / "logo.jpg"
REPORT_COLORS = ["#F29129", "#FFC46B", "#F7A94C", "#FFD994", "#E68A22", "#FFE7B8", "#CC741C", "#FFF0CF", "#B86212", "#F6D08A"]
CHART_GRADIENT_START = colors.HexColor("#EEF6FF")
CHART_GRADIENT_MIDDLE = colors.HexColor("#FBFDFF")
CHART_GRADIENT_END = colors.HexColor("#D8E5F0")


def safe_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def generate_xml(notas, filtros, output):
    """
    Gera relatório XML das notas fiscais cadastradas.
    XML completo conforme os campos disponíveis no banco do projeto.
    """

    root = Element("relatorio_notas_fiscais")

    metadata = SubElement(root, "metadata")
    SubElement(metadata, "sistema").text = "NFE Scanner"
    SubElement(metadata, "data_geracao").text = datetime.now().isoformat()
    SubElement(metadata, "filtros").text = safe_text(filtros)
    SubElement(metadata, "total_notas").text = str(len(notas))
    SubElement(metadata, "valor_total").text = str(sum(n.valor_total or 0 for n in notas))

    notas_node = SubElement(root, "notas")

    for nota in notas:
        nota_node = SubElement(notas_node, "nota_fiscal")

        SubElement(nota_node, "id").text = safe_text(nota.id)
        SubElement(nota_node, "numero_nf").text = safe_text(nota.numero_nf)
        SubElement(nota_node, "serie").text = safe_text(nota.serie)
        SubElement(nota_node, "data_emissao").text = safe_text(nota.data_emissao)
        SubElement(nota_node, "cnpj_fornecedor").text = safe_text(nota.cnpj_fornecedor)
        SubElement(nota_node, "nome_fornecedor").text = safe_text(nota.nome_fornecedor)
        SubElement(nota_node, "valor_total").text = safe_text(nota.valor_total)
        SubElement(nota_node, "chave_acesso").text = safe_text(nota.chave_acesso)
        SubElement(nota_node, "local").text = safe_text(nota.local)
        SubElement(nota_node, "produto").text = safe_text(nota.produto)
        SubElement(nota_node, "quantidade").text = safe_text(nota.quantidade)
        SubElement(nota_node, "transportador").text = safe_text(nota.transportador)
        SubElement(nota_node, "faturista").text = safe_text(nota.faturista)
        SubElement(nota_node, "lider_operacional").text = safe_text(nota.lider_operacional)
        SubElement(nota_node, "observacao").text = safe_text(nota.observacao)
        SubElement(nota_node, "caminho_arquivo_imagem").text = safe_text(nota.caminho_arquivo_imagem)
        SubElement(nota_node, "data_cadastro").text = safe_text(nota.data_cadastro)

    tree = ElementTree(root)
    indent(tree, space="    ", level=0)

    tree.write(output, encoding="utf-8", xml_declaration=True)


def format_ton(value):
    return f"{float(value or 0):,.3f}".replace(",", "X").replace(".", ",").replace("X", ".") + " TON"


def format_period(inicio, fim):
    return f"{inicio.strftime('%d/%m/%Y %H:%M')} ate {fim.strftime('%d/%m/%Y %H:%M')}"


def _blend_color(start, end, ratio):
    return colors.Color(
        start.red + (end.red - start.red) * ratio,
        start.green + (end.green - start.green) * ratio,
        start.blue + (end.blue - start.blue) * ratio,
    )


def _add_chart_gradient(drawing, width, height):
    steps = 48
    stripe_width = width / steps
    for index in range(steps):
        position = index / max(steps - 1, 1)
        if position <= 0.46:
            fill = _blend_color(CHART_GRADIENT_START, CHART_GRADIENT_MIDDLE, position / 0.46)
        else:
            fill = _blend_color(CHART_GRADIENT_MIDDLE, CHART_GRADIENT_END, (position - 0.46) / 0.54)
        drawing.add(Rect(index * stripe_width, 0, stripe_width + 1, height, strokeColor=fill, fillColor=fill))
    drawing.add(Rect(0, 0, width, height, strokeColor=colors.HexColor("#C6D6E2"), fillColor=None, strokeWidth=0.8))


def _pdf_table(rows, widths=None, header_rows=1):
    table = Table(rows, colWidths=widths, repeatRows=header_rows)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, header_rows - 1), BRAND_ORANGE),
                ("TEXTCOLOR", (0, 0), (-1, header_rows - 1), colors.white),
                ("FONTNAME", (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, BRAND_LINE),
                ("ROWBACKGROUNDS", (0, header_rows), (-1, -1), [colors.white, BRAND_LIGHT]),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def _section_heading(title, subtitle=None):
    content = [Paragraph(f"<b>{title}</b>", getSampleStyleSheet()["Heading2"])]
    if subtitle:
        content.append(Paragraph(subtitle, getSampleStyleSheet()["Normal"]))
    section = Table([[content]], colWidths=[250 * mm])
    section.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), BRAND_LIGHT),
                ("BOX", (0, 0), (-1, -1), 0.8, BRAND_LINE),
                ("LINEBEFORE", (0, 0), (0, -1), 4, BRAND_ORANGE),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return section


def _chart_card(drawing):
    card = Table([[drawing]], colWidths=[250 * mm])
    card.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF6FF")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#C6D6E2")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return card


def _draw_pdf_branding(canvas, doc):
    page_width, page_height = landscape(A4)
    canvas.saveState()
    canvas.setStrokeColor(BRAND_LINE)
    canvas.setLineWidth(0.8)
    canvas.roundRect(8 * mm, 8 * mm, page_width - 16 * mm, page_height - 16 * mm, 3 * mm, stroke=1, fill=0)
    canvas.setFillColor(BRAND_ORANGE)
    canvas.rect(8 * mm, page_height - 11 * mm, page_width - 16 * mm, 3 * mm, stroke=0, fill=1)
    if LOGO_PATH.exists():
        canvas.drawImage(str(LOGO_PATH), 15 * mm, page_height - 31 * mm, width=18 * mm, height=18 * mm, preserveAspectRatio=True, mask="auto")
    canvas.setFillColor(BRAND_DARK)
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawString(37 * mm, page_height - 21 * mm, "SCAN-NFE MINASFALTO")
    canvas.setFillColor(BRAND_MUTED)
    canvas.setFont("Helvetica", 8)
    canvas.drawString(37 * mm, page_height - 26 * mm, "Relatorio operacional de notas fiscais")
    canvas.setStrokeColor(BRAND_LINE)
    canvas.line(15 * mm, page_height - 34 * mm, page_width - 15 * mm, page_height - 34 * mm)
    canvas.line(15 * mm, 18 * mm, page_width - 15 * mm, 18 * mm)
    canvas.setFillColor(BRAND_MUTED)
    canvas.setFont("Helvetica", 7)
    canvas.drawString(15 * mm, 13 * mm, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    canvas.drawRightString(page_width - 15 * mm, 13 * mm, f"Pagina {doc.page}")
    canvas.restoreState()


def _pie_drawing(periodo, title):
    drawing = Drawing(720, 235)
    _add_chart_gradient(drawing, 720, 235)
    drawing.add(String(360, 222, title, textAnchor="middle", fontName="Helvetica-Bold", fontSize=13))
    produtos = [item for item in periodo["produtos"] if item["quantidade_ton"] > 0]
    if not produtos:
        drawing.add(String(360, 110, "Nenhum dado no periodo", textAnchor="middle", fontSize=10))
        return drawing

    pie = Pie()
    pie.x = 275
    pie.y = 20
    pie.width = 180
    pie.height = 180
    pie.data = [item["quantidade_ton"] for item in produtos]
    pie.labels = [f"{item['produto']} ({format_ton(item['quantidade_ton'])})" for item in produtos]
    pie.slices.strokeWidth = 0.5
    pie.slices.strokeColor = colors.white
    for index, _produto in enumerate(produtos):
        pie.slices[index].fillColor = colors.HexColor(REPORT_COLORS[index % len(REPORT_COLORS)])
    pie.sideLabels = True
    drawing.add(pie)
    return drawing


def _receipt_bar_drawing(recebimento):
    drawing = Drawing(720, 270)
    _add_chart_gradient(drawing, 720, 270)
    title = "Recebimento diario"
    if recebimento["material"]:
        title += f" - {recebimento['material']}"
    drawing.add(String(360, 255, title, textAnchor="middle", fontName="Helvetica-Bold", fontSize=13))
    materiais = [item["material"] for item in recebimento["totais_materiais"]]
    if not materiais:
        drawing.add(String(360, 130, "Nenhum dado no periodo", textAnchor="middle", fontSize=10))
        return drawing

    chart = VerticalBarChart()
    chart.x = 55
    chart.y = 45
    chart.width = 610
    chart.height = 175
    chart.data = [[dia["materiais_ton"].get(material, 0) for dia in recebimento["dias"]] for material in materiais]
    chart.categoryAxis.categoryNames = [item["data"][8:10] for item in recebimento["dias"]]
    chart.categoryAxis.style = "stacked"
    chart.categoryAxis.labels.fontSize = 7
    chart.valueAxis.valueMin = 0
    for index, _material in enumerate(materiais):
        chart.bars[index].fillColor = colors.HexColor(REPORT_COLORS[index % len(REPORT_COLORS)])
    chart.barLabelFormat = [
        ["%0.3f" if value >= 8 else None for value in serie]
        for serie in chart.data
    ]
    chart.barLabels.nudge = 0
    chart.barLabels.boxTarget = "mid"
    chart.barLabels.boxAnchor = "c"
    chart.barLabels.textAnchor = "middle"
    chart.barLabels.fontName = "Helvetica-Bold"
    chart.barLabels.fontSize = 6
    chart.barLabels.fillColor = BRAND_DARK
    drawing.add(chart)
    legend_x = max(50, 360 - len(materiais) * 42)
    for index, material in enumerate(materiais):
        drawing.add(String(legend_x + index * 84, 20, material[:16], fillColor=colors.HexColor(REPORT_COLORS[index % len(REPORT_COLORS)]), fontName="Helvetica-Bold", fontSize=7))
    return drawing


def _receipt_share_drawing(recebimento):
    drawing = Drawing(720, 230)
    _add_chart_gradient(drawing, 720, 230)
    drawing.add(String(360, 215, "Participacao total por material", textAnchor="middle", fontName="Helvetica-Bold", fontSize=13))
    materiais = recebimento["totais_materiais"]
    if not materiais:
        drawing.add(String(360, 110, "Nenhum dado no periodo", textAnchor="middle", fontSize=10))
        return drawing

    pie = Pie()
    pie.x = 260
    pie.y = 25
    pie.width = 200
    pie.height = 165
    pie.data = [item["total_ton"] for item in materiais]
    pie.labels = [f"{item['material']} ({format_ton(item['total_ton'])})" for item in materiais]
    for index, _material in enumerate(materiais):
        pie.slices[index].fillColor = colors.HexColor(REPORT_COLORS[index % len(REPORT_COLORS)])
        pie.slices[index].strokeColor = colors.white
    pie.sideLabels = True
    pie.simpleLabels = False
    drawing.add(pie)
    return drawing


def _period_from_material(material):
    return {
        "inicio": material["inicio"],
        "fim": material["fim"],
        "total_ton": material["total_ton"],
        "total_notas": material["total_nfes"],
        "produtos": [
            {"produto": item["material"], "quantidade_ton": item["quantidade_ton"]}
            for item in material["materiais"]
        ],
    }


def generate_operational_pdf(operacional, material, setor, recebimento, output):
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=36 * mm,
        bottomMargin=24 * mm,
        title="SCAN-NFE MINASFALTO - Relatorio Operacional",
        author="SCAN-NFE MINASFALTO",
    )

    periodo_filtrado = _period_from_material(material)
    summary_rows = [["Periodo", "Intervalo", "Total", "Quantidade NF-es"]]
    summary_rows.append([
        "Periodo filtrado",
        format_period(periodo_filtrado["inicio"], periodo_filtrado["fim"]),
        format_ton(periodo_filtrado["total_ton"]),
        periodo_filtrado["total_notas"],
    ])
    material_rows = [["Material", "Quantidade Produto", "Quantidade NF-es"]]
    material_rows.extend([[item["material"], format_ton(item["quantidade_ton"]), item["quantidade_nfes"]] for item in material["materiais"]])
    setor_rows = [["Material", "Produto CDMA", "NF CDMA", "Produto PRU", "NF PRU"]]
    setor_rows.extend(
        [
            [
                item["material"],
                format_ton(item["quantidade_cdma_ton"]),
                item["quantidade_nfes_cdma"],
                format_ton(item["quantidade_pru_ton"]),
                item["quantidade_nfes_pru"],
            ]
            for item in setor["materiais"]
        ]
    )
    story = [
        _section_heading("Resumo operacional", "Visao consolidada dos recebimentos conforme o filtro aplicado."),
        Spacer(1, 10),
        _pdf_table(summary_rows, [35 * mm, 95 * mm, 40 * mm, 40 * mm]),
        PageBreak(),
        _section_heading("Acumulado do periodo", format_period(periodo_filtrado["inicio"], periodo_filtrado["fim"])),
        Spacer(1, 10),
        _chart_card(_pie_drawing(periodo_filtrado, "Quantidade por produto")),
        PageBreak(),
        _section_heading("Relatorio por Periodo e por Material", format_period(material["inicio"], material["fim"])),
        Spacer(1, 10),
        _pdf_table(material_rows, [130 * mm, 55 * mm, 45 * mm]),
        Spacer(1, 16),
        _section_heading("Relatorio por Periodo, Material e Local", format_period(setor["inicio"], setor["fim"])),
        Spacer(1, 10),
        _pdf_table(setor_rows, [90 * mm, 45 * mm, 30 * mm, 45 * mm, 30 * mm]),
        PageBreak(),
        _section_heading("Quantidade e Proporcao de Materiais por Dia", format_period(recebimento["inicio"], recebimento["fim"])),
        Spacer(1, 10),
        _chart_card(_receipt_bar_drawing(recebimento)),
        PageBreak(),
        _section_heading("Participacao total por material", f"Material: {recebimento['material'] or 'Todos os materiais'}"),
        Spacer(1, 10),
        _chart_card(_receipt_share_drawing(recebimento)),
    ]
    doc.build(story, onFirstPage=_draw_pdf_branding, onLaterPages=_draw_pdf_branding)


def _style_sheet(sheet, header_row=1):
    orange = PatternFill("solid", fgColor="F29129")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D8D3CA")
    for cell in sheet[header_row]:
        cell.fill = orange
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in sheet.iter_rows(min_row=header_row):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")
    sheet["A1"].font = Font(size=16, bold=True, color="B85F00")
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 3, 55)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def generate_operational_excel(operacional, material, setor, recebimento, output):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Relatorio Operacional NF-e"])
    summary.append([f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    summary.append([])
    periodo_filtrado = _period_from_material(material)
    summary.append(["Periodo", "Inicio", "Fim", "Total TON", "Quantidade NF-es"])
    summary.append([
        "Periodo filtrado",
        periodo_filtrado["inicio"],
        periodo_filtrado["fim"],
        periodo_filtrado["total_ton"],
        periodo_filtrado["total_notas"],
    ])
    summary.append([])
    summary.append(["Produto periodo", "Quantidade TON"])
    for item in periodo_filtrado["produtos"]:
        summary.append([item["produto"], item["quantidade_ton"]])
    _style_sheet(summary, 4)

    material_sheet = workbook.create_sheet("Materiais")
    material_sheet.append(["Relatorio por Periodo e por Material"])
    material_sheet.append([format_period(material["inicio"], material["fim"])])
    material_sheet.append([])
    material_sheet.append(["Material", "Quantidade Produto TON", "Quantidade NF-es"])
    for item in material["materiais"]:
        material_sheet.append([item["material"], item["quantidade_ton"], item["quantidade_nfes"]])
    _style_sheet(material_sheet, 4)

    setor_sheet = workbook.create_sheet("Materiais por Local")
    setor_sheet.append(["Relatorio por Periodo, Material e Local"])
    setor_sheet.append([format_period(setor["inicio"], setor["fim"])])
    setor_sheet.append([])
    setor_sheet.append(["Material", "Produto CDMA TON", "NF CDMA", "Produto PRU TON", "NF PRU"])
    for item in setor["materiais"]:
        setor_sheet.append(
            [
                item["material"],
                item["quantidade_cdma_ton"],
                item["quantidade_nfes_cdma"],
                item["quantidade_pru_ton"],
                item["quantidade_nfes_pru"],
            ]
        )
    _style_sheet(setor_sheet, 4)

    receipt_sheet = workbook.create_sheet("Recebimento Diario")
    receipt_sheet.append(["Quantidade e Proporcao de Materiais por Dia"])
    receipt_sheet.append([format_period(recebimento["inicio"], recebimento["fim"])])
    receipt_sheet.append([f"Material: {recebimento['material'] or 'Todos os materiais'}"])
    materiais_recebimento = [item["material"] for item in recebimento["totais_materiais"]]
    receipt_sheet.append(["Data", *[f"{material} TON" for material in materiais_recebimento]])
    for item in recebimento["dias"]:
        receipt_sheet.append([item["data"], *[item["materiais_ton"].get(material, 0) for material in materiais_recebimento]])
    _style_sheet(receipt_sheet, 4)
    receipt_chart = BarChart()
    receipt_chart.type = "col"
    receipt_chart.style = 10
    receipt_chart.title = "Quantidade por Dia, Material e Local"
    receipt_chart.y_axis.title = "Toneladas"
    receipt_chart.x_axis.title = "Data de emissao"
    receipt_chart.height = 12
    receipt_chart.width = 24
    receipt_chart.grouping = "stacked"
    receipt_chart.overlap = 100
    receipt_chart.add_data(Reference(receipt_sheet, min_col=2, max_col=receipt_sheet.max_column, min_row=4, max_row=receipt_sheet.max_row), titles_from_data=True)
    receipt_chart.set_categories(Reference(receipt_sheet, min_col=1, min_row=5, max_row=receipt_sheet.max_row))
    receipt_chart.dLbls = DataLabelList()
    receipt_chart.dLbls.showVal = True
    receipt_chart.dLbls.numFmt = "0.000"
    receipt_sheet.add_chart(receipt_chart, "E4")
    workbook.save(output)
